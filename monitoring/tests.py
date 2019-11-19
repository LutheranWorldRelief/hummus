"""
Tests for 'monitoring'
"""
import datetime
import json
import time

from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.gis.geos import Point
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.test import TestCase, Client
from django.utils import translation, formats

from constance import config
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from monitoring.models import Contact, Sex, Template
from monitoring.updates import update_contact
from monitoring.common import parse_date, xstr


class ImportTestCase(TestCase):

    language = translation.get_supported_language_variant(translation.get_language())
    start_row = config.START_ROW
    header_row = config.HEADER_ROW
    template = config.DEFAULT_TEMPLATE
    date_format = formats.get_format("SHORT_DATE_FORMAT")
    params = {'language': language, 'start_row': start_row, 'header_row': header_row,
              'template': template, 'date_format': date_format, }

    def setUp(self):

        # creates template
        clean_template = Template()
        clean_template.id = self.template
        clean_template.name = 'Test Template'
        clean_template.mapping = json.loads('{"contact": {"sex": {"name": "Sexo", "varname": "sex", "required": false}, "country": {"name": "País", "varname": "country", "required": true}, "document": {"name": "Número de Identificación", "varname": "document", "required": false}, "men_home": {"name": "Hombres en su familia", "varname": "men_home", "required": false}, "birthdate": {"name": "Fecha de Nacimiento", "varname": "birthdate", "required": false}, "community": {"name": "Comunidad", "varname": "community", "required": false}, "education": {"name": "Educación", "varname": "education", "required": false}, "last_name": {"name": "Apellidos", "varname": "last_name", "required": false}, "first_name": {"name": "Nombres", "varname": "first_name", "required": true}, "women_home": {"name": "Mujeres en su familia", "varname": "women_home", "required": false}, "municipality": {"name": "Departamento", "varname": "municipality", "required": false}, "organization": {"name": "Organización Perteneciente", "varname": "organization", "required": false}, "phone_personal": {"name": "Teléfono", "varname": "phone_personal", "required": false}}, "project": {"name": {"name": "Nombre de Proyecto", "varname": "name", "required": true}, "organization": {"name": "Organización Implementadora", "varname": "organization", "required": true}}, "project_contact": {"date_entry_project": {"name": "Fecha de ingreso al proyecto", "varname": "date_entry_project", "required": true}}}')
        clean_template.save()

        # creates excel file
        excel_name = 'test_excel_file.xlsx'
        tmp_excel_name = "{}-{}-{}".format('test', time.strftime("%Y%m%d-%H%M%S"),
                                           excel_name)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'data'
        self.params['excel_file'] = tmp_excel_name

        # creates excel content
        row = 1
        column = 1
        sheet.cell(row=row, column=column).value = 'Merged titles'
        row += 1
        for model in clean_template.mapping:
            for field_name, field_details in clean_template.mapping[model].items():
                sheet.cell(row=row, column=column).value = field_details['name']
                column += 1
        row += 1
        default_storage.save('{}/{}'.
                             format('tmp', tmp_excel_name), ContentFile(save_virtual_workbook(workbook)))

    def test_step3(self):
        admin = User.objects.create_superuser('admin', email=None, password=None)
        c = Client()
        c.force_login(admin)
        response = c.post('/import/participants/step3', self.params)
        self.assertEqual(response.status_code, 200)


class CommonTestCase(TestCase):

    def test_xstr(self):
        self.assertEqual('', xstr(None))
        self.assertEqual('', xstr(''))
        self.assertEqual('Rosita Oconnor', xstr('Rosita Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('Rosita  Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('Rosita   Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('Rosita Oconnor '))
        self.assertEqual('Rosita Oconnor', xstr('Rosita Oconnor   '))
        self.assertEqual('Rosita Oconnor', xstr(' Rosita Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('  Rosita Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('   Rosita Oconnor'))
        self.assertEqual('Rosita Oconnor', xstr('   Rosita    Oconnor   '))

    def test_parse_date(self):
        test_date = datetime.date(2000, 12, 31)
        test_datetime = datetime.datetime(2000, 12, 31, 0, 0, 0, 0)
        self.assertEqual(test_date, parse_date('2000-12-31 00:00:00'))
        self.assertEqual(test_date, parse_date(test_datetime))
        self.assertEqual(test_date, parse_date(test_date))
        self.assertEqual(test_date, parse_date('2000-12-31'))
        self.assertEqual(test_date, parse_date('12/31/2000'))
        self.assertEqual(None, parse_date(1))

class ContactTestCase(TestCase):

    point = Point(1, 1)
    name = "Ophelia Oshiro"
    document = "555"
    contact1 = None
    contact2 = None

    def setUp(self):
        point = self.point
        name = self.name
        document = self.document

        # Supporting tables
        Sex.objects.create(name='Man', varname='man', id='M')
        Sex.objects.create(name='Woman', varname='woman', id='F')

        # dupe
        self.contact1 = Contact.objects.create(name=name, document=document, location=point)
        self.contact2 = Contact.objects.create(name=name, document=document, location=point)

        # other entries
        Contact.objects.create(first_name="Trey", last_name="Polk", location=point)
        Contact.objects.create(first_name="Gilberto", last_name="Gilsdorf")

    def test_name_generation(self):
        """name is created from first and last name"""
        contact = Contact.objects.create(first_name="Maranda", last_name="Pedro")
        self.assertEqual(contact.name, "Maranda Pedro")
        Contact.objects.create(name="Bruce Lehr")
        contact = Contact.objects.get(name="Bruce Lehr")
        self.assertEqual(contact.name, "Bruce Lehr")

    def test_snake_names(self):
        """changes snake_name to Snake Name"""
        Contact.objects.create(first_name="Audry", last_name="Glassford", community='snake_name')
        contact = Contact.objects.get(first_name="Audry", last_name="Glassford")
        self.assertEqual(contact.community, "Snake Name")

    def test_remove_spaces(self):
        """trims name"""
        contact = Contact.objects.create(first_name=" Willena  ", last_name="   Vanalstyne ")
        self.assertEqual(contact.name, "Willena Vanalstyne")
        self.assertEqual(contact.first_name, "Willena")
        self.assertEqual(contact.last_name, "Vanalstyne")

    def test_no_name(self):
        """tries to add a contact with no name"""
        with self.assertRaisesMessage(ValueError, 'We need a name! name, first_name and last_name seem to be empty.'):
            Contact.objects.create(document='555')

    def test_update_contact(self):
        """add a contact looking up FK using names """
        request = None
        row = {}
        row['first_name'] = 'Diana'
        row['last_name'] = 'Lacey'
        row['sex'] = 'Woman'
        contact = Contact()
        update_contact(request, contact, row)
        self.assertEqual(contact.name, "Diana Lacey")
        self.assertEqual(contact.sex.id, "F")

    def test_update_varname(self):
        """add a contact looking up FK using varnames """
        request = None
        row = {}
        row['name'] = 'Amado Oberholtzer'
        row['sex'] = 'man'
        contact = Contact()
        update_contact(request, contact, row)
        self.assertEqual(contact.name, "Amado Oberholtzer")
        self.assertEqual(contact.sex.id, "M")

    def test_find_dupes(self):
        """find duplicates using names"""
        name = self.name
        c = Client()
        response = c.get('/opt/api-name/{}/'.format(name.upper()))
        self.assertEqual(response.status_code, 200)
        response_content = json.loads(response.content)
        self.assertEqual(name, response_content['models'][0]['name'])

    def test_api_contact(self):
        """get contact json data using id"""
        name = self.name
        document = self.document
        contact = Contact.objects.filter(name=name, document=document).first()
        c = Client()
        response = c.get('/opt/api-contact/{}/'.format(contact.id))
        self.assertEqual(response.status_code, 200)
        response_content = json.loads(response.content)
        self.assertEqual(name, response_content['models'][0]['name'])

    def test_api_doc(self):
        """get contact json data using document"""
        name = self.name
        document = self.document
        c = Client()
        response = c.get('/opt/api-doc/{}/'.format(document))
        self.assertEqual(response.status_code, 200)
        response_content = json.loads(response.content)
        self.assertEqual(name, response_content['models'][0]['name'])

    def test_api_doc_values(self):
        """get contact json data details"""
        name = self.name
        document = self.document
        id1 = self.contact1.id
        c = Client()
        response = c.post('/opt/api-doc-values/'.format(document), {'ids[]': id1})
        self.assertEqual(response.status_code, 200)
        response_content = json.loads(response.content)
        self.assertEqual(name, response_content['values']['name'])
