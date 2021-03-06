"""
create custom clean template with data validation
"""
from django.utils.translation import gettext_lazy as _

from constance import config
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from .models import SubProject, Sex, Education, Country, Product, ContactType
from .common import get_localized_name as __


# creates catalog and validations
def create_catalog(book, request):

    dvs = {}
    sheet = book.create_sheet(__('catalog'))
    catalog_cols = (SubProject, Sex, Education, Country, Product, ContactType)
    col_start = 1
    row_start = config.HEADER_ROW
    for col, col_value in enumerate(catalog_cols, start=col_start):
        sheet.cell(row=row_start, column=col, value=col_value.__name__)
    row_start = config.START_ROW
    for col, col_value in enumerate(catalog_cols, start=col_start):
        if hasattr(col_value.objects, 'for_user'):
            rows = col_value.objects.for_user(request.user).all()
        else:
            rows = col_value.objects.all()
        row = 0
        for row, row_value in enumerate(rows, start=row_start):
            name = __('name') if hasattr(row_value, __('name')) else 'name'
            sheet.cell(row=row, column=col, value=getattr(row_value, name))
        letter = get_column_letter(col)
        dvs[col_value.__name__] = DataValidation(
            type="list", allow_blank=True, showDropDown=False,
            formula1="%s!$%s$2:$%s$%s" % (__('catalog'), letter, letter, row), )
        col += 1


    # applies validations
    sheet = book[_('data')]
    # TODO this should be mapped for each template differently
    validation_cols = {'A': 'SubProject', 'E': 'Sex', 'G': 'Education', 'L': 'Country', }
    row_start = config.START_ROW
    row_max = 2000

    for letter in validation_cols:
        data_validation = dvs[validation_cols[letter]]
        col = column_index_from_string(letter)
        sheet.add_data_validation(data_validation)
        data_validation.add('%s%s:%s%s' % (letter, row_start, letter, row_max))

    return sheet
