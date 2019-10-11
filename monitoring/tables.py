"""
django-tables2 and django-filter views or definitions
"""
from os.path import basename

from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse

from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Submit
from django_filters import FilterSet, CharFilter, ModelChoiceFilter
from django_select2.forms import Select2Widget
from django_tables2 import Table, SingleTableView, RequestConfig, Column
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from .models import Template, Project, Country, Organization, Contact, ProjectContact, SubProject
from .common import get_localized_name as __
from .catalog import create_catalog


class ReportExportMixin:
    export_name = "table"
    export_trigger_param = "_export"
    exclude_columns = ()

    def table_to_dataset(self, table, exclude_columns):
        dataset = []
        for row_num, row in enumerate(table.as_values(exclude_columns=exclude_columns)):
            if row_num == 0:
                continue
            dataset.append(row)
        return dataset

    def get_export_filename(self, export_format):
        return "{}.{}".format(self.export_name, export_format)

    def create_export(self, export_format):

        if export_format != 'xlsx':
            raise ValueError(_('Unsupported export format.'))

        # get table and create dataset
        table = self.get_table(**self.get_table_kwargs())
        dataset = self.table_to_dataset(table, self.exclude_columns)

        # get localized excel tempalte
        obj = Template.objects.get(id='clean-template')
        tfile = getattr(obj, __('file'))
        tfilename = tfile.name

        # loads 'datos' sheet # TODO rename to 'data'?
        book = load_workbook(filename=tfile)
        create_catalog(book, self.request)
        sheet = book[_('data')]

        # adds rows
        cur = 3  # Force start at row 3
        for row, row_entry in enumerate(dataset, start=1):
            for col, col_entry in enumerate(row_entry, start=1):
                sheet.cell(row=row + cur - 1, column=col, value=col_entry)

        # resposne
        response = HttpResponse(content=save_virtual_workbook(book),
                                content_type='application/vnd.openxmlformats-officedocument.'
                                'spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % (basename(tfilename),)
        return response

    def render_to_response(self, context, **kwargs):
        export_format = self.request.GET.get(self.export_trigger_param, None)
        if export_format == "xlsx":
            return self.create_export(export_format)

        return super().render_to_response(context, **kwargs)


class PagedFilteredTableView(SingleTableView):
    filter_class = None
    formhelper_class = None
    context_filter_name = 'filter'

    def get_queryset(self):
        if hasattr(self.model.objects, 'for_user'):
            qs = self.model.objects.for_user(self.request.user)
        else:
            qs = super(PagedFilteredTableView, self).get_queryset()

        self.filter = self.filter_class(self.request.GET, request=self.request, queryset=qs)
        self.filter.form.helper = self.formhelper_class()
        return self.filter.qs

    def get_table(self, **kwargs):
        table = super(PagedFilteredTableView, self).get_table()
        if 'page' in self.kwargs:
            page = self.kwargs['page']
        else:
            page = 1
        RequestConfig(self.request, paginate={'page': page,
                                              "per_page": self.paginate_by}).configure(table)
        return table

    def get_context_data(self, **kwargs):
        context = super(PagedFilteredTableView, self).get_context_data()
        context[self.context_filter_name] = self.filter
        context['export_formats'] = ('xlsx',)
        return context


#
# Specific tables
#


# ProjectContact

class ProjectContactTable(Table):

    sex = Column(accessor="contact.sex.name")
    education = Column(accessor="contact.education.name")

    class Meta:
        model = ProjectContact
        attrs = {'class': 'table table-bordered', 'id': 'tbPreview'}
        fields = ('project', 'organization', 'contact.document', 'contact.first_name',
                  'contact.last_name', 'sex', 'contact.birthdate',
                  'education', 'contact.phone', 'contact.men', 'contact.women',
                  'contact.organization', 'contact.country.name', 'contact.deparment',
                  'contact.community', 'contact.startdate', 'product')

    def render_education(self, value, record):
        return getattr(record.contact.education, __('name'))

    def render_sex(self, value, record):
        return getattr(record.contact.sex, __('name'))


def projects(request):
    if request is None:
        return Project.objects.none()
    return Project.objects.for_user(request.user).all()


def countries(request):
    if request is None:
        return Country.objects.none()
    return Country.objects.for_user(request.user).all()


def organizations(request):
    if request is None:
        return Organization.objects.none()
    return Organization.objects.for_user(request.user).all()


class ProjectContactFilter(FilterSet):
    contact__name = CharFilter(lookup_expr='icontains')
    project = ModelChoiceFilter(queryset=projects, widget=Select2Widget)
    contact__country = ModelChoiceFilter(
        queryset=countries, widget=Select2Widget)
    organization = ModelChoiceFilter(
        queryset=organizations, widget=Select2Widget)

    class Meta:
        model = ProjectContact
        fields = ('contact__country', 'project',
                  'organization', 'contact__name')


class ProjectContactFilterFormHelper(FormHelper):
    form_method = 'GET'
    layout = Layout(
        'contact__country',
        'project',
        'organization',
        'contact__name',
        Submit('submit', 'Apply Filter'),
    )


# Contact

class ContactTable(Table):
    def render_name(self, value, record):
        url = record.get_absolute_url()
        return mark_safe('<a href="%s">%s</a>' % (url, value))

    class Meta:
        model = Contact
        fields = ('name', 'country', 'sex')


class ContactFilter(FilterSet):
    class Meta:
        model = Contact
        fields = ('name', 'country', 'sex')


class ContactFilterFormHelper(FormHelper):
    form_method = 'GET'
    layout = Layout(
        'name',
        'country',
        'sex',
        Submit('submit', 'Apply Filter'),
    )


# Project

class ProjectTable(Table):
    def render_name(self, value, record):
        url = record.get_absolute_url()
        return mark_safe('<a href="%s">%s</a>' % (url, value))

    class Meta:
        model = Project
        fields = ('name',)


class ProjectFilter(FilterSet):
    name = CharFilter(lookup_expr='icontains')

    class Meta:
        fields = ('name',)
        model = Project


class ProjectFilterFormHelper(FormHelper):
    form_method = 'GET'
    layout = Layout(
        'name',
        Submit('submit', 'Apply Filter'),
    )


# SubProject

class SubProjectTable(ProjectTable):
    class Meta:
        fields = ('name',)
        model = SubProject


class SubProjectFilter(ProjectFilter):
    class Meta:
        fields = ('name',)
        model = SubProject


class SubProjectFilterFormHelper(ProjectFilterFormHelper):
    pass
