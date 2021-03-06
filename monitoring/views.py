"""
'monitoring' views, mostly invoked by urls.py
"""
import json
import time
from os.path import basename

from django.conf import settings
from django.db.models import Q, Value, F
from django.db.models.functions import Upper, Trim, Coalesce
from django.contrib.gis.geos import Point
from django.core.files.storage import default_storage
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.utils import translation, formats
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import TemplateView, DetailView, FormView
from django.views.decorators.csrf import csrf_exempt

from constance import config
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from .tables import (SubProjectTable, ProjectTable, ContactTable, ProjectContactTable,
                     PagedFilteredTableView, ReportExportMixin,
                     SubProjectFilter, SubProjectFilterFormHelper,
                     ProjectFilter, ProjectFilterFormHelper,
                     ProjectContactFilter, ProjectContactFilterFormHelper,
                     ContactFilter, ContactFilterFormHelper, )
from .models import (SubProject, Project, Contact, Template, Organization, ProjectContact,
                     Request, City, Profile, Log, Country)
from .common import (DomainRequiredMixin, MONTHS, get_localized_name as __,
                     RegexpReplace, parse_date, xstr)
from .catalog import create_catalog
from .updates import update_contact, update_project_contact, validate_data, try_to_find


class GetExcelToImport(DomainRequiredMixin, View):
    """
    Also known as 'step1'.
    Receives the excel file from user and sets advanced options.
    """

    template_name = 'import/step1.html'

    def get(self, request):
        context = {}
        languages = [{'value': short_name,
                      'name': long_name} for (short_name, long_name) in
                     Profile.LANGUAGE_CHOICES]
        templates = Template.objects.values(template_id=F('id'),
                                            template_name=F(__('name')))
        context['short_date_format'] = formats.get_format("SHORT_DATE_FORMAT")
        context['languages'] = languages
        context['templates'] = templates
        return render(request, self.template_name, context)


class ValidateExcel(DomainRequiredMixin, FormView):
    """
    Also known as 'step2'.
    Validates the excel file and gets user confirmation to import.
    """

    def post(self, request, *args, **kwargs):
        messages_error = []

        # get advanced options
        language = request.POST.get('language', translation.get_supported_language_variant(
            settings.LANGUAGE_CODE))
        start_row = int(request.POST.get('start_row', config.START_ROW))
        header_row = int(request.POST.get('header_row', config.HEADER_ROW))
        template = request.POST.get('template', config.DEFAULT_TEMPLATE)
        date_format = request.POST.get('date_format', formats.get_format("SHORT_DATE_FORMAT"))

        # get file and set name
        excel_file = request.FILES['excel_file']
        tmp_excel_name = "{}-{}-{}".format(request.user.username, time.strftime("%Y%m%d-%H%M%S"),
                                           excel_file.name)
        default_storage.save('tmp/{}'.format(tmp_excel_name), excel_file)
        uploaded_wb = load_workbook(filename=excel_file)

        # gets the data sheet, tries all languages
        uploaded_ws = None
        if _('data') in uploaded_wb.sheetnames:
            uploaded_ws = uploaded_wb[_('data')]
        else:
            for other_lang in [lang[0] for lang in settings.LANGUAGES]:
                with translation.override(other_lang):
                    if _('data') in uploaded_wb.sheetnames:
                        uploaded_ws = uploaded_wb[_('data')]
                        continue
        if not uploaded_ws:
            uploaded_ws = uploaded_wb.active

        # check headers
        template_obj = Template.objects.get(id=template)
        mapping = getattr(template_obj, __('mapping', language))
        columns_required = []
        headers = [cell.value for cell in uploaded_ws[header_row]]

        # checks columns from mapping exist in uploaded file
        for model, fields in mapping.items():
            for field_name, field_data in fields.items():
                column_header = field_data['name']

                if column_header not in headers:
                    raise Exception('Column "{}" not found, choices are: {}'
                                    .format(column_header, ', '.join(filter(None, headers))))

                if field_data['required']:
                    columns_required.append(column_header)

        context = {}
        context['columns'] = uploaded_ws[header_row]

        headers = {cell.value: cell.col_idx - 1 for cell in uploaded_ws[header_row]}
        uploaded_ws.delete_rows(0, amount=start_row - 1)

        # map headers to columns
        for model, fields in mapping.items():
            for field_name, field_data in fields.items():
                column_header = field_data['name']
                mapping[model][field_name]['column'] = headers[column_header]

        # validate rows
        for row in uploaded_ws.iter_rows():
            # skips empty row
            if xstr(row[0].value) == '' and xstr(row[1].value) == '' and xstr(row[2].value) == '':
                continue
            # quick data validation
            error_message = validate_data(row, mapping, start_row, date_format, language)
            if error_message:
                messages_error.append(error_message)

        grouped_errors = {}
        for row in messages_error:
            row_messages = row['msgs']
            for row_message in row_messages:
                reference, clean_msg = row_message.split(': ', 1)
                if clean_msg not in grouped_errors:
                    grouped_errors[clean_msg] = {}
                    grouped_errors[clean_msg]['reference'] = []
                    grouped_errors[clean_msg]['count'] = 0
                grouped_errors[clean_msg]['count'] += 1
                grouped_errors[clean_msg]['reference'].append(reference)

        grouped_errors = sorted(grouped_errors.items(),
                                key=lambda k_v: k_v[1]['count'], reverse=True)

        excel_data_rows = uploaded_ws
        if uploaded_ws.max_row > 20:
            excel_data_rows = uploaded_ws.iter_rows(min_row=1, max_row=20)

        context['grouped_errors'] = grouped_errors
        context['data'] = excel_data_rows
        context['columns_required'] = columns_required
        context['start_row'] = start_row
        context['date_format'] = date_format
        context['excel_file'] = tmp_excel_name
        context['template'] = template
        context['header_row'] = header_row
        context['language'] = language
        context['sheet'] = uploaded_ws.title

        return render(request, self.template_name, context)

    template_name = 'import/step2.html'


class ImportParticipants(DomainRequiredMixin, FormView):
    """
    Also known as 'step3'.
    Further validates rows and proceeds to import.
    """

    def post(self, request, *args, **kwargs):
        messages_error = []
        messages_info = []
        imported_ids = []
        counter_records_updated = 0
        counter_records_created = 0
        filter_type = 'iexact'

        # get advanced options
        language = request.POST.get('language', translation.get_supported_language_variant(
            settings.LANGUAGE_CODE))
        start_row = int(request.POST.get('start_row', config.START_ROW))
        header_row = int(request.POST.get('header_row', config.HEADER_ROW))
        template = request.POST.get('template', config.DEFAULT_TEMPLATE)
        date_format = request.POST.get('date_format', formats.get_format("SHORT_DATE_FORMAT"))

        # read excel file and sheet
        sheet = request.POST.get('sheet', _('data'))
        tmp_excel = request.POST.get('excel_file')
        excel_file = default_storage.open('{}/{}'.format('tmp', tmp_excel))
        uploaded_wb = load_workbook(excel_file)
        uploaded_ws = uploaded_wb[sheet]

        # check headers
        template_obj = Template.objects.get(id=template)
        headers = {cell.value: cell.col_idx - 1 for cell in uploaded_ws[header_row]}
        mapping = getattr(template_obj, __('mapping', language))

        # map headers to columns
        for model, fields in mapping.items():
            for field_name, field_data in fields.items():
                column_header = field_data['name']
                mapping[model][field_name]['column'] = headers[column_header]

        # create log event
        content = "Excel import using '{}' and template '{}'".format(tmp_excel, template)
        log = Log.objects.create(module='excel import', user=request.user.username, content=content)

        # import
        uploaded_ws.delete_rows(0, amount=start_row - 1)
        for row in uploaded_ws.iter_rows():
            # skips empty row, based on first 3 values
            if xstr(row[0].value) == '' and xstr(row[1].value) == '' and xstr(row[2].value) == '':
                continue

            # quick data validation
            error_message = validate_data(row, mapping)
            if error_message:
                messages_error.append(error_message)
                continue

            # create or update contact
            model = Contact
            model_fields = mapping['contact']
            row_dict = {}
            row_dict['source_id'] = 'excel'
            for field_name, field_data in model_fields.items():
                value = row[field_data['column']].value
                if value:
                    if model._meta.get_field(field_name).get_internal_type() == 'PointField':
                        (lat, lng) = value.split(' ')[:2]
                        value = Point(float(lng), float(lat))
                    if model._meta.get_field(field_name).get_internal_type() == 'CharField':
                        value = str(value)
                    if model._meta.get_field(field_name).get_internal_type() == 'DateField':
                        if not row[field_data['column']].is_date:
                            value = parse_date(value, date_format)
                else:
                    value = None

                # removes extra spaces if string
                if isinstance(value, str):
                    value = xstr(value)

                row_dict[field_name] = value

            # there are two ways to look up a contact: name+doc and firstname+lastname+doc
            if {'name', 'document'} <= set(model_fields):
                contact = Contact.objects.filter(name__iexact=row_dict['name'],
                                                 document__iexact=row_dict['document)']).first()
            elif {'first_name', 'last_name', 'document'} <= set(model_fields):
                name = "{} {}".format(xstr(row_dict['first_name']), xstr(row_dict['last_name']))
                row_dict['name'] = xstr(name)
                contact = Contact.objects.filter(Q(name__iexact=row_dict['name'],
                                                   document__iexact=row_dict['document']) |
                                                 Q(first_name__iexact=row_dict['name'],
                                                   last_name__iexact=row_dict['last_name'],
                                                   document__iexact=row_dict['document'])).first()
            else:
                raise Exception('Mapping needs more contact data fields')

            # create new organization if needed
            contact_organization = Organization.objects.filter(
                name__iexact=row_dict['organization']).first()
            if not contact_organization and row_dict['organization']:
                contact_organization = Organization.objects.filter(
                    varname__iexact=row_dict['organization']).first()
            if not contact_organization and row_dict['organization']:
                messages_info.append('Create organization: {}'.format(row_dict['organization']))
                contact_organization = Organization()
                contact_organization.name = row_dict['organization']
                contact_organization.created_user = request.user.username
                contact_organization.log = log
                contact_organization.save()

            # create contact if needed
            if not contact:
                messages_info.append('Create contact: {}'.format(contact))
                contact = Contact()
                if contact_organization:
                    contact.organization = contact_organization
                counter_records_created += 1
            else:
                messages_info.append('Update contact: {}'.format(contact))
                counter_records_updated += 1
            row_dict['log'] = log
            update_contact(request, contact, row_dict)

            imported_ids.append(contact.id)

            # create or update project contact
            model = ProjectContact
            model_fields = mapping['project_contact']
            row_dict = {}
            row_dict['source_id'] = 'excel'

            for field_name, field_data in model_fields.items():
                value = row[field_data['column']].value
                # removes extra spaces if string
                if isinstance(value, str):
                    value = xstr(value)
                if value:
                    if model._meta.get_field(field_name).get_internal_type() == 'ForeignKey':
                        related_model = model._meta.get_field(field_name).related_model
                        value = try_to_find(related_model, value)
                    if model._meta.get_field(field_name).get_internal_type() == 'DateField':
                        if not row[field_data['column']].is_date:
                            value = parse_date(value, date_format)
                else:
                    value = None
                row_dict[field_name] = value

            subproject = row_dict['subproject']
            project_contact = ProjectContact.objects.filter(subproject=subproject,
                                                            contact=contact).first()
            if not project_contact:
                messages_info.append('Create participant: {} {}'.format(subproject, contact))
                project_contact = ProjectContact()
                project_contact.contact = contact
            else:
                messages_info.append('Update participant: {} {}'.format(subproject, contact))
            row_dict['log'] = log
            update_project_contact(request, project_contact, row_dict)

        # gets dupes
        contacts = Contact.objects.all()
        if self.request.user:
            contacts = contacts.for_user(self.request.user)
        contacts = contacts.annotate(
            name_uc=Trim(Upper(RegexpReplace(F('name'), r'\s+', ' ', 'g'))))

        # manually (python) counts dupes, because count messed up the distinct() filter
        names = {}
        for row in contacts:
            if row.name_uc not in names:
                names[row.name_uc] = 0
            names[row.name_uc] += 1
        names_uc = list(k_v[0] for k_v in names.items() if k_v[1] > 1)

        contacts_names_ids = contacts.values_list('id', flat=True).filter(name_uc__in=names_uc)

        # manually (python) counts dupes, because count messed up the distinct() filter
        contacts = contacts.filter(document__isnull=False).exclude(document='')
        docs = {}
        for row in contacts:
            if row.document not in docs:
                docs[row.document] = 0
            docs[row.document] += 1
        documents = list(k_v[0] for k_v in docs.items() if k_v[1] > 1)

        contacts = Contact.objects.filter(id__in=imported_ids) \
            .filter(Q(id__in=contacts_names_ids) | Q(document__in=documents)) \
            .values(contact_id=F('id'),
                    contact_name=Coalesce('name', Value('')),
                    contact_sex=Coalesce('sex_id', Value('')),
                    contact_document=Coalesce('document', Value('')),
                    contact_organization=Coalesce('organization__name', Value('')),
                    )

        context = {}
        context['excel_file'] = tmp_excel
        context['messages_error'] = messages_error
        context['messages_info'] = messages_info
        context['quantity_records_updated'] = counter_records_updated
        context['quantity_records_created'] = counter_records_created
        context['model'] = list(contacts)
        return render(request, self.template_name, context)

    template_name = 'import/step3.html'


@method_decorator(csrf_exempt, name='dispatch')
class Capture(TemplateView):
    template_name = 'capture.html'

    # TODO: GET should be removed
    def get(self, request, *args, **kwargs):
        context = {}
        context['meta'] = request.META
        context['body'] = request.body.decode('utf-8')
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        context = {}
        row = Request()
        row.meta = request.META
        row.body = request.body.decode('utf-8')
        body = json.loads(row.body)
        row_dict = {}
        project_name = body['form'].get('project')
        project = Project.objects.get(name=project_name)
        try:
            row_dict['name'] = body['form'].get('name', '')
            row_dict['first_name'] = body['form'].get('first_name', '')
            row_dict['last_name'] = body['form'].get('last_name', '')
            row_dict['name'] = body['form'].get('name', '')
            row_dict['sex'] = body['form'].get('sex', '')
            row_dict['country'] = body['form'].get('country', '')
            row_dict['education'] = body['form'].get('education', '')
            row_dict['document'] = body['form'].get('id', '')
            row_dict['source_id'] = 'commcare'
        except KeyError as e:
            print('KeyError in data forwarding : "%s"' % str(e))

        # try to find contact
        contact = Contact.objects.filter(document__iexact=row_dict['document'],
                                         first_name__iexact=row_dict['first_name'],
                                         last_name__iexact=row_dict['last_name']).first()

        # using MDC sometimes only 'name' is collected, try to find contact
        if not contact:
            contact = Contact.objects.filter(document__iexact=row_dict['document'],
                                             name__iexact=row_dict['name']).first()

        if not contact:
            print('Create contact: {} {} {}'.format(row_dict['name'],
                                                    row_dict['first_name'], row_dict['last_name']))
            contact = Contact()
            update_contact(request, contact, row_dict)
        else:
            print('Update contact: {} {} {}'.format(row_dict['name'],
                                                    row_dict['first_name'], row_dict['last_name']))
            update_contact(request, contact, row_dict)

        project_contact = ProjectContact.objects.filter(project__name__iexact=project_name,
                                                        contact=contact).first()
        if not project_contact:
            print('Create project contact: {} {}'.format(project.name, row_dict['name']))
            project_contact = ProjectContact()
            project_contact.contact = contact
            project_contact.project = project
            update_project_contact(request, project_contact, row_dict)
        else:
            print('Update project contact: {} {}'.format(project.name, row_dict['first_name']))
            update_project_contact(request, project_contact, row_dict)
        return render(request, self.template_name, context)


class DownloadTemplate(DomainRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # get localized excel template
        obj = Template.objects.get(id=config.DEFAULT_TEMPLATE)
        tfile = getattr(obj, __('file'))
        tfilename = tfile.name

        book = load_workbook(filename=tfile)
        create_catalog(book, request)

        # response
        response = HttpResponse(content=save_virtual_workbook(book),
                                content_type='application/vnd.openxmlformats-officedocument.'
                                             'spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % (basename(tfilename),)
        return response


class ValidateDupesDoc(DomainRequiredMixin, TemplateView):
    template_name = 'dupes_document.html'


class ValidateDupesName(DomainRequiredMixin, TemplateView):
    template_name = 'dupes_name.html'


class ValidateDupesNameFuzzy(DomainRequiredMixin, TemplateView):
    template_name = 'dupes_name_fuzzy.html'


class SubProjectTableView(DomainRequiredMixin, PagedFilteredTableView):
    model = SubProject
    table_class = SubProjectTable
    template_name = 'table.html'
    paginate_by = 50
    filter_class = SubProjectFilter
    formhelper_class = SubProjectFilterFormHelper


class ProjectTableView(DomainRequiredMixin, PagedFilteredTableView):
    model = Project
    table_class = ProjectTable
    template_name = 'table.html'
    paginate_by = 50
    filter_class = ProjectFilter
    formhelper_class = ProjectFilterFormHelper


class ContactTableView(DomainRequiredMixin, PagedFilteredTableView):
    model = Contact
    table_class = ContactTable
    template_name = 'table.html'
    paginate_by = 50
    filter_class = ContactFilter
    formhelper_class = ContactFilterFormHelper


class ProjectContactTableView(DomainRequiredMixin, ReportExportMixin, PagedFilteredTableView):
    model = ProjectContact
    table_class = ProjectContactTable
    template_name = 'table.html'
    paginate_by = 50
    filter_class = ProjectContactFilter
    formhelper_class = ProjectContactFilterFormHelper


class DashboardView(DomainRequiredMixin, TemplateView):
    template_name = 'modular_template/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)

        # sets custom user filters in query string
        user = self.request.user
        query_string = []
        if hasattr(user, 'profile'):
            project_ids = user.profile.projects.values_list('pk', flat=True)
            for project_id in project_ids:
                query_string.append('project_ids[]={}'.format(project_id))
            if len(project_ids) == 1:
                query_string.append('project_id={}'.format(project_ids[0]))
            for country_id in user.profile.countries.values_list('pk', flat=True):
                query_string.append('country_id[]={}'.format(country_id))
            for lwrregion_id in user.profile.lwrregions.values_list('pk', flat=True):
                query_string.append('country_id[]={}'.format(lwrregion_id))
            if len(query_string) > 0:
                query_string.append('extra_counters=1')
                query_string.append('my_dashboard=true')
        context['query_string'] = '&'.join(query_string)

        years = Project.objects.order_by('start__fyear').exclude(projectcontact__isnull=True). \
            values_list('start__fyear', flat=True).distinct()
        context['months'] = MONTHS
        context['years'] = list(years)

        if self.request.GET.get('project_id'):
            project_id = self.request.GET.get('project_id')
            context['project'] = Project.objects.filter(id=project_id).values('id', 'name').first()

        if self.request.GET.get('subproject_id'):
            subproject_id = self.request.GET.get('subproject_id')
            context['subproject'] = SubProject.objects.filter(
                id=subproject_id).values('id', 'name').first()

        if self.request.GET.getlist('country_id[]'):
            countries = self.request.GET.getlist('country_id[]')
            context['countries'] = list(Country.objects.filter(
                id__in=countries).values('id', 'name'))

        if self.request.GET.getlist('lwrregion_id[]'):
            context['regions'] = self.request.GET.getlist('lwrregion_id[]')

        if self.request.GET.getlist('year[]'):
            context['years_data'] = self.request.GET.getlist('year[]')

        if self.request.GET.get('quarter'):
            context['quarter'] = self.request.GET.get('quarter')

        context['my_dashboard'] = self.request.GET.get('my_dashboard', False)

        return context


class SubProjectDetailView(DetailView):
    model = SubProject

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contact_locations'] = self.object.project.projectcontact_set.filter(
            contact__location__isnull=False)
        return context


class ProjectDetailView(DetailView):
    model = Project

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contact_locations'] = self.object.projectcontact_set.filter(
            contact__location__isnull=False)
        return context


class ContactDetailView(DetailView):
    model = Contact

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class CityDetailView(DetailView):
    model = City

    def get_object(self, queryset=None):
        name = self.kwargs.get('name')
        country_id = self.kwargs.get('country_id')
        obj = City.objects.all()
        if country_id:
            obj = obj.filter(Q(country_id__iexact=country_id.upper()) |
                             Q(country__name__iexact=country_id))
        obj = obj.filter(name__iexact=name).first()
        return obj
